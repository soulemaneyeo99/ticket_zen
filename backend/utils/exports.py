"""
Service d'export de données (CSV, Excel, PDF)
"""
import csv
import io
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


class ExportService:
    """Service pour exporter des données"""
    
    def export_data(self, export_type, export_format='csv', date_from=None, date_to=None, user=None):
        """
        Exporter des données selon le type et le format
        
        Args:
            export_type: Type de données à exporter (tickets, payments, trips, etc.)
            export_format: Format d'export (csv, excel, pdf)
            date_from: Date de début (optionnel)
            date_to: Date de fin (optionnel)
            user: Utilisateur qui exporte (pour filtrage)
        
        Returns:
            dict: Informations sur le fichier exporté
        """
        # Récupérer les données selon le type
        data = self._get_data(export_type, date_from, date_to, user)
        
        # Exporter selon le format
        if export_format == 'csv':
            return self._export_csv(data, export_type)
        elif export_format == 'excel':
            return self._export_excel(data, export_type)
        elif export_format == 'pdf':
            return self._export_pdf(data, export_type)
        else:
            raise ValueError(f'Format non supporté: {export_format}')
    
    def _get_data(self, export_type, date_from, date_to, user):
        """Récupérer les données à exporter"""
        from apps.tickets.models import Ticket
        from apps.payments.models import Payment
        from apps.trips.models import Trip
        from apps.companies.models import Company
        
        queryset = None
        
        if export_type == 'tickets':
            queryset = Ticket.objects.select_related(
                'trip', 'trip__company', 'passenger'
            ).all()
            
            if user and user.role == 'compagnie':
                queryset = queryset.filter(trip__company=user.company)
            
            if date_from:
                queryset = queryset.filter(created_at__gte=date_from)
            if date_to:
                queryset = queryset.filter(created_at__lte=date_to)
            
            return self._format_tickets_data(queryset)
        
        elif export_type == 'payments':
            queryset = Payment.objects.select_related(
                'user', 'trip', 'company'
            ).all()
            
            if user and user.role == 'compagnie':
                queryset = queryset.filter(company=user.company)
            
            if date_from:
                queryset = queryset.filter(created_at__gte=date_from)
            if date_to:
                queryset = queryset.filter(created_at__lte=date_to)
            
            return self._format_payments_data(queryset)
        
        elif export_type == 'trips':
            queryset = Trip.objects.select_related(
                'company', 'vehicle', 'departure_city', 'arrival_city'
            ).all()
            
            if user and user.role == 'compagnie':
                queryset = queryset.filter(company=user.company)
            
            if date_from:
                queryset = queryset.filter(departure_datetime__gte=date_from)
            if date_to:
                queryset = queryset.filter(departure_datetime__lte=date_to)
            
            return self._format_trips_data(queryset)
        
        elif export_type == 'companies':
            queryset = Company.objects.all()
            
            if date_from:
                queryset = queryset.filter(created_at__gte=date_from)
            if date_to:
                queryset = queryset.filter(created_at__lte=date_to)
            
            return self._format_companies_data(queryset)
        
        else:
            raise ValueError(f'Type d\'export non supporté: {export_type}')
    
    def _format_tickets_data(self, queryset):
        """Formater les données des tickets"""
        headers = [
            'Numéro Ticket', 'Passager', 'Email', 'Téléphone',
            'Départ', 'Arrivée', 'Date Départ', 'Siège',
            'Prix', 'Statut', 'Payé', 'Date Création'
        ]
        
        rows = []
        for ticket in queryset:
            rows.append([
                ticket.ticket_number,
                ticket.passenger_full_name,
                ticket.passenger_email,
                ticket.passenger_phone,
                ticket.trip.departure_city.name,
                ticket.trip.arrival_city.name,
                ticket.trip.departure_datetime.strftime('%d/%m/%Y %H:%M'),
                ticket.seat_number,
                f'{ticket.total_amount} FCFA',
                ticket.get_status_display(),
                'Oui' if ticket.is_paid else 'Non',
                ticket.created_at.strftime('%d/%m/%Y %H:%M')
            ])
        
        return {'headers': headers, 'rows': rows, 'title': 'Export Tickets'}
    
    def _format_payments_data(self, queryset):
        """Formater les données des paiements"""
        headers = [
            'Transaction ID', 'Utilisateur', 'Montant', 'Méthode',
            'Statut', 'Commission', 'Montant Compagnie',
            'Date', 'Complété Le'
        ]
        
        rows = []
        for payment in queryset:
            rows.append([
                payment.transaction_id,
                payment.user.get_full_name(),
                f'{payment.amount} FCFA',
                payment.get_payment_method_display(),
                payment.get_status_display(),
                f'{payment.platform_commission} FCFA',
                f'{payment.company_amount} FCFA',
                payment.created_at.strftime('%d/%m/%Y %H:%M'),
                payment.completed_at.strftime('%d/%m/%Y %H:%M') if payment.completed_at else 'N/A'
            ])
        
        return {'headers': headers, 'rows': rows, 'title': 'Export Paiements'}
    
    def _format_trips_data(self, queryset):
        """Formater les données des voyages"""
        headers = [
            'Compagnie', 'Départ', 'Arrivée', 'Date Départ',
            'Prix', 'Places Totales', 'Places Disponibles',
            'Taux Occupation', 'Statut', 'Revenu'
        ]
        
        rows = []
        for trip in queryset:
            rows.append([
                trip.company.name,
                trip.departure_city.name,
                trip.arrival_city.name,
                trip.departure_datetime.strftime('%d/%m/%Y %H:%M'),
                f'{trip.base_price} FCFA',
                trip.total_seats,
                trip.available_seats,
                f'{trip.occupancy_rate:.1f}%',
                trip.get_status_display(),
                f'{trip.total_revenue} FCFA'
            ])
        
        return {'headers': headers, 'rows': rows, 'title': 'Export Voyages'}
    
    def _format_companies_data(self, queryset):
        """Formater les données des compagnies"""
        headers = [
            'Nom', 'Email', 'Téléphone', 'Ville', 'Statut',
            'Total Voyages', 'Total Tickets', 'Revenu Total',
            'Date Création'
        ]
        
        rows = []
        for company in queryset:
            rows.append([
                company.name,
                company.email,
                company.phone_number,
                company.city,
                company.get_status_display(),
                company.total_trips,
                company.total_tickets_sold,
                f'{company.total_revenue} FCFA',
                company.created_at.strftime('%d/%m/%Y')
            ])
        
        return {'headers': headers, 'rows': rows, 'title': 'Export Compagnies'}
    
    def _export_csv(self, data, export_type):
        """Exporter en CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Écrire les headers
        writer.writerow(data['headers'])
        
        # Écrire les données
        for row in data['rows']:
            writer.writerow(row)
        
        # Créer le nom de fichier
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'export_{export_type}_{timestamp}.csv'
        
        # Sauvegarder le fichier
        from django.conf import settings
        import os
        
        file_path = os.path.join(settings.MEDIA_ROOT, 'exports', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
            f.write(output.getvalue())
        
        file_url = f'{settings.MEDIA_URL}exports/{filename}'
        
        return {
            'filename': filename,
            'file_url': file_url,
            'format': 'csv',
            'rows_count': len(data['rows'])
        }
    
    def _export_excel(self, data, export_type):
        """Exporter en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = data['title']
        
        # Style pour les headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Écrire les headers
        for col_num, header in enumerate(data['headers'], 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Écrire les données
        for row_num, row_data in enumerate(data['rows'], 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder le fichier
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'export_{export_type}_{timestamp}.xlsx'
        
        from django.conf import settings
        import os
        
        file_path = os.path.join(settings.MEDIA_ROOT, 'exports', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        wb.save(file_path)
        
        file_url = f'{settings.MEDIA_URL}exports/{filename}'
        
        return {
            'filename': filename,
            'file_url': file_url,
            'format': 'excel',
            'rows_count': len(data['rows'])
        }
    
    def _export_pdf(self, data, export_type):
        """Exporter en PDF"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'export_{export_type}_{timestamp}.pdf'
        
        from django.conf import settings
        import os
        
        file_path = os.path.join(settings.MEDIA_ROOT, 'exports', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Créer le PDF
        doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        
        # Titre
        title = Paragraph(data['title'], title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.5*inch))
        
        # Tableau
        table_data = [data['headers']] + data['rows']
        table = Table(table_data)
        
        # Style du tableau
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        file_url = f'{settings.MEDIA_URL}exports/{filename}'
        
        return {
            'filename': filename,
            'file_url': file_url,
            'format': 'pdf',
            'rows_count': len(data['rows'])
        }